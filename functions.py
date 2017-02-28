import os, shutil, pprint
from openpyxl import load_workbook, worksheet
from openpyxl.utils import get_column_letter
import sys
import time

# TODO - implement ability to deal with multi-length filenames

def choose_working_dir():
    print('Work in the current directory?\n\n{}'
          .format(os.path.dirname(os.path.realpath(__file__))))
    if yes_or_no():
        return str(os.path.dirname(os.path.realpath(__file__)))
    else:
        path = str(input('Please enter the full path to the directory'
                         ' you want to work in \n\n'))
        if not path.endswith('\\'):
            return "{}\\".format(path)
        else:
            return path


def split_file_names(directory):
    '''
    split .wav file names into lists and add them to a list
    :return list of lists:
    '''
    file_name_list_split = []
    for i in os.listdir(directory):
        if not str(i).endswith('.wav'):
            continue
        # Split filename into a list.
        file_name_list_split.append(i.split('_'))
    return file_name_list_split


def get_file_names_raw(directory):
    '''
    returns a list of .wav filestring names
    :param directory:
    :return:
    '''
    file_name_list_raw = []
    for i in os.listdir(directory):
        if not str(i).endswith('.wav'):
            continue
        file_name_list_raw.append(i)
    return file_name_list_raw


# TODO - Make this work!
def tally_alts(split_list):
    '''
    iterate over list of lists, placing the first 4 indeces into a dictionary as a key with value 1
    and incrementing that value per identical entry
    :return:
    '''
    filename_dict = {}
    for i in split_list:
        if '{}_{}_{}_{}'.format(i[0], i[1], i[2], i[3]) not in filename_dict:
            filename_dict['{}_{}_{}_{}'.format(i[0], i[1], i[2], i[3])] = 1
        else:
            filename_dict['{}_{}_{}_{}'.format(i[0], i[1], i[2], i[3])] += 1


def checkMissing(script_file_list, file_names):
    '''
    compare dicts script and files, returns a with missing values
    :param script_file_list:
    dictionary
    :param file_names:
    dictionary
    :return:
    '''
    missing_files = {}
    for i in script_file_list:
        if i + '.wav' not in file_names:
            missing_files[i] = True
    return missing_files


def getWorkbook(directory):
    '''
    iterate over a directory to find .xlsx files and adds them to a list.
    :return:list
    :param directory:
    '''
    workbook_names = []
    for i in os.listdir(directory):
        if not str(i).endswith('.xlsx'):
            continue
        workbook_names.append(i)
    if not len(workbook_names):
        return False
    elif len(workbook_names) == 1:
        return str(workbook_names[0])
    else:
        return workbook_names


def choose_workbook(directory):
    '''
    looks in directory and adds found xlsx files
    to a list which is returned only if there is more than one entry.
    otherwise exit()
    :param directory:
    :return: string or exit()
    '''

    workbook_names = []
    workbook_names.append(getWorkbook(directory))
    if False in workbook_names:
        print('No .xlsx files found. Exiting...')
        exit()
    if len(workbook_names) == 1:
        return str(workbook_names[0])
    if len(workbook_names) > 1:
        # TODO - allow user to choose a file from this list and return it
        print("There are multiple .xlsx files in the directory:")
        pprint.pprint(workbook_names)
        exit()


def yes_or_no():
    yes = set(['yes', 'y', 'ye', ''])
    no = set(['no', 'n'])

    choice = input().lower()
    if choice in yes:
        return True
    elif choice in no:
        return False
    else:
        sys.stdout.write("Please respond with 'yes' or 'no'\n")


def open_workbook(path):
    '''
    opens target path as a workbook object
    :param path:
    :return: wb object
    '''


def choose_wb_sheet(wb):
    '''
    prints names of sheets in a workbook object along with their index in a list.
    user input gives index of chosen sheet
    returns string of the name of sheet chosen by user
    :param wb: workbook object
    :return: string of chosen sheet from wb
    '''
    ws_list = wb.get_sheet_names()
    for i in ws_list:
        print('{}. {}'.format(ws_list.index(i), str(i)))
    while True:
        try:
            sheet_number = int(input("Which # sheet do you want to open?\n"))
        except ValueError:
            pass
        if (sheet_number > -1) and (sheet_number <= len(ws_list) - 1):
            break
    return str(ws_list[sheet_number])


def find_ws_column(ws, keyword, range):
    '''
    iterate rows in ws, searching for string 'KEYWORD'
    :param ws: ws object
    :param keyword: string to search for in column
    :param range: range string of worksheet
    :return: index of column that begins with "keyword"
    '''
    for row in ws.iter_rows(range):
        for cell in row:
            if cell.value == keyword:
                return cell.column


def get_script_names(ws, column):
    '''
    loops over a column, discarding empty cells and creates a list containing the values
    :param ws: worksheet containing column
    :param column: column as int to be looped
    #:param list_name: list to add values to
    :return:
    '''
    list_name = []
    column_letter = get_column_letter(column)
    for row in ws.iter_rows('{}{}:{}{}'.format(str(column_letter),
                                               str(ws.min_row),
                                               str(column_letter),
                                               str(ws.max_row))):
        for cell in row:
            if cell.value:
                list_name.append(cell.value)
    return list_name


def get_time_stamp():
    timestr = time.strftime("%Y%m%d-%H%M%S")
    return str("\n\n{}\n".format(timestr))


def gameloft_builder(master_folder, working_dir):
    """
    iterate over wav files in a directory, and place them in gameloft build file structure
    e.g. 'M1151_INTRO_01_VAHN.wav'
    :param master_folder: string - Name of parent folder created for file structure
    :param working_dir: directory path - directory to work in
    :return: nothing
    """
    for i in os.listdir(working_dir):
        if not str(i).endswith('.wav'):
            continue
        # Split filename into a list.
        file_name_list = i.split('_')
        # file_path = os.path.abspath(i)
        file_path = str(working_dir + i)
        # Use first 2 indexes in fileNameList to create correct path for file.
        dest_path = ('.\\' + str(master_folder) + '\\' + str(file_name_list[0]) + '\\' + str(file_name_list[1]))
        # Copy file to correct path in file structure.
        if not os.path.exists(dest_path):
            os.makedirs(dest_path)
        shutil.copy(file_path, dest_path)
    print('you did it champ!!')


def main():
    dir_to_use = choose_working_dir()
    file_name_list = get_file_names_raw(dir_to_use)
    print('Got {} files!'.format(len(file_name_list)))
    wb = load_workbook(str(dir_to_use) + "//" +
                       str(choose_workbook(dir_to_use)),
                       data_only=True, read_only=True)

    ws = wb.get_sheet_by_name(choose_wb_sheet(wb))

    keyword = str(input('Enter the name of the column '
                        'header you are looking for (e.g. FILENAME):\n'))

    script_file_list = get_script_names(ws, (find_ws_column(ws, keyword, 'A1:Q3')))

    missing_files = checkMissing(script_file_list, file_name_list)
    print('The following files are missing:')
    if missing_files:
        for key in missing_files:
            if key == 'FILENAME':
                continue
            print(key)
        with open('Missing-files.txt', 'a') as text_file:
            text_file.write(get_time_stamp())
        for key in missing_files:
            if key == 'FILENAME':
                continue
            with open('Missing-files.txt', 'a') as text_file:
                text_file.write('{}\n'.format(key))
    print('Missing filenames written to Missing-files.txt')

    sys.stdout.write('Begin sorting files?\n')
    if yes_or_no():
        gameloft_builder(str(input('Enter master folder name \n e.g. "DH5_GU15"\n')), dir_to_use)
    else:
        print('Bye Bye!')

if __name__ == '__main__':
    main()
